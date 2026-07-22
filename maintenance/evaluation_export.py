import csv
import json
import re
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from statistics import mean
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font

from accounts.ml.maintenance_nlp import classify_issue_category, predict_priority


ROOT_DIR = Path(__file__).resolve().parents[1]
DEFAULT_OWNER_WORKBOOK = ROOT_DIR / "exports" / "ml" / "category_evaluation_reviewer.xlsx"
DEFAULT_SOURCE_DATASET = ROOT_DIR / "exports" / "complete_dataset" / "maintenance_requests.csv"
DEFAULT_REPORTS_DIR = ROOT_DIR / "reports"

CATEGORY_ORDER = ["PLUMBING", "ELECTRICAL", "STRUCTURAL", "OTHER"]
PRIORITY_ORDER = ["LOW", "MEDIUM", "HIGH"]

CATEGORY_DISPLAY = {
    "PLUMBING": "Plumbing",
    "ELECTRICAL": "Electrical",
    "STRUCTURAL": "Structural",
    "OTHER": "Other",
}
PRIORITY_DISPLAY = {
    "LOW": "Low",
    "MEDIUM": "Medium",
    "HIGH": "High",
}

CSV_COLUMNS = [
    "record_code",
    "sanitized_title",
    "sanitized_description",
    "human_category",
    "system_category",
    "category_correct",
    "keyword_matched",
    "human_priority",
    "system_priority",
    "priority_confidence",
    "priority_correct",
]

EMAIL_RE = re.compile(r"\b[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}\b", re.IGNORECASE)
PHONE_RE = re.compile(
    r"(?:(?<!\d)(?:\+?63|0)\d{10}(?!\d)|(?<!\d)\d{3}[- ]?\d{3}[- ]?\d{4}(?!\d)|(?<!\d)\d{4}[- ]?\d{3}[- ]?\d{4}(?!\d))"
)
UNIT_RE = re.compile(r"\b(unit)\s*#?\s*[A-Z0-9-]+\b", re.IGNORECASE)
ACCOUNT_RE = re.compile(
    r"\b(?:acct(?:ount)?|account|payment|gcash|maya|card|reference|ref(?:erence)?)\b[:#\s-]*[A-Z0-9-]{4,}",
    re.IGNORECASE,
)
LONG_DIGIT_RE = re.compile(r"(?<!\d)\d{4,19}(?!\d)")
USERNAME_RE = re.compile(r"\b(username|user|tenant|name)\s*[:=-]\s*[A-Z0-9._@-]+", re.IGNORECASE)
WHITESPACE_RE = re.compile(r"\s+")

EXPECTED_CATEGORY_METRICS = {
    "total_records": 197,
    "correct": 164,
    "incorrect": 33,
    "accuracy": 0.8325,
    "macro_precision": 0.8605,
    "macro_recall": 0.8641,
    "macro_f1": 0.8358,
    "keyword_match_coverage": 0.8274,
    "unmatched_rate": 0.1726,
}
EXPECTED_PRIORITY_METRICS = {
    "total_records": 197,
    "accuracy": 0.3249,
    "macro_f1": 0.2847,
    "human_distribution": {"LOW": 9, "MEDIUM": 80, "HIGH": 108},
}


@dataclass(frozen=True)
class ExportRecord:
    record_code: str
    test_id: str
    sanitized_title: str
    sanitized_description: str
    human_category_raw: str
    human_category_normalized: str
    system_category: str
    category_correct: bool
    keyword_matched: bool
    human_priority_raw: str
    human_priority_normalized: str
    system_priority: str
    priority_confidence: float
    priority_correct: bool


def sanitize_text(text: str) -> str:
    value = text or ""
    value = EMAIL_RE.sub("[Redacted Email]", value)
    value = PHONE_RE.sub("[Redacted Phone]", value)

    def replace_unit(match):
        word = match.group(1)
        return f"{word.title()} [Redacted]"

    value = UNIT_RE.sub(replace_unit, value)
    value = ACCOUNT_RE.sub("[Redacted Payment Info]", value)
    value = LONG_DIGIT_RE.sub("[Redacted Number]", value)
    value = USERNAME_RE.sub(lambda m: f"{m.group(1)}: [Redacted]", value)
    return WHITESPACE_RE.sub(" ", value).strip()


def normalize_category_label(raw_value: str) -> str:
    normalized = (raw_value or "").strip().upper()
    if normalized not in CATEGORY_ORDER:
        raise ValueError(f"Unsupported category label: {raw_value!r}")
    return normalized


def normalize_priority_label(raw_value: str) -> str:
    normalized = (raw_value or "").strip().upper()
    if normalized not in PRIORITY_ORDER:
        raise ValueError(f"Unsupported priority label: {raw_value!r}")
    return normalized


def title_case_category(normalized: str) -> str:
    return CATEGORY_DISPLAY[normalized]


def title_case_priority(normalized: str) -> str:
    return PRIORITY_DISPLAY[normalized]


def normalize_test_id(raw_test_id: str) -> str:
    value = (raw_test_id or "").strip()
    if value.upper().startswith("MR-"):
        value = value[3:].strip()
    return value


def read_source_dataset(source_csv_path: Path) -> dict[str, dict]:
    if not source_csv_path.exists():
        raise FileNotFoundError(f"Source dataset not found: {source_csv_path}")
    with source_csv_path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        rows = list(reader)
    data = {}
    for row in rows:
        test_id = (row.get("id") or "").strip()
        if test_id:
            data[test_id] = row
    return data


def find_required_columns(header_values: list[str]) -> dict[str, int]:
    normalized = {str(value).strip().lower(): index for index, value in enumerate(header_values)}
    required = {
        "test_id": "test_id",
        "maintenance_description": "maintenance_description",
        "human_label": "human_label",
        "priority_level": "priority_level",
    }
    column_map = {}
    missing = []
    for key, header in required.items():
        if header not in normalized:
            missing.append(header)
        else:
            column_map[key] = normalized[header]
    if missing:
        raise ValueError(f"Owner-labeled workbook is missing required columns: {', '.join(missing)}")
    return column_map


def read_owner_labeled_rows(workbook_path: Path) -> list[dict]:
    if not workbook_path.exists():
        raise FileNotFoundError(f"Owner-labeled workbook not found: {workbook_path}")
    workbook = load_workbook(workbook_path, data_only=False)
    worksheet = workbook[workbook.sheetnames[0]]
    header_values = [worksheet.cell(1, column).value for column in range(1, worksheet.max_column + 1)]
    column_map = find_required_columns(header_values)

    rows = []
    for row_number in range(2, worksheet.max_row + 1):
        values = [worksheet.cell(row_number, column).value for column in range(1, worksheet.max_column + 1)]
        if all(value in (None, "") for value in values):
            continue
        rows.append(
            {
                "excel_row": row_number,
                "raw_test_id": "" if values[column_map["test_id"]] is None else str(values[column_map["test_id"]]),
                "maintenance_description": ""
                if values[column_map["maintenance_description"]] is None
                else str(values[column_map["maintenance_description"]]),
                "human_label": "" if values[column_map["human_label"]] is None else str(values[column_map["human_label"]]),
                "priority_level": ""
                if values[column_map["priority_level"]] is None
                else str(values[column_map["priority_level"]]),
            }
        )
    return rows


def load_owner_label_records(
    workbook_path: Path = DEFAULT_OWNER_WORKBOOK,
    source_csv_path: Path = DEFAULT_SOURCE_DATASET,
) -> list[dict]:
    source_data = read_source_dataset(source_csv_path)
    workbook_rows = read_owner_labeled_rows(workbook_path)
    if not workbook_rows:
        raise ValueError(f"No owner-labeled rows found in workbook: {workbook_path}")

    seen_ids = set()
    records = []
    errors = []
    for row in workbook_rows:
        test_id = normalize_test_id(row["raw_test_id"])
        if not test_id:
            errors.append(f"Excel row {row['excel_row']}: blank test_id")
            continue
        if test_id in seen_ids:
            errors.append(f"Excel row {row['excel_row']}: duplicate test_id '{test_id}'")
            continue
        seen_ids.add(test_id)
        if test_id not in source_data:
            errors.append(f"Excel row {row['excel_row']}: test_id '{test_id}' not found in source dataset")
            continue

        source_row = source_data[test_id]
        source_description = source_row.get("description") or ""
        if row["maintenance_description"] != source_description:
            errors.append(f"Excel row {row['excel_row']}: description mismatch for test_id '{test_id}'")
            continue

        try:
            human_category = normalize_category_label(row["human_label"])
        except ValueError as exc:
            errors.append(f"Excel row {row['excel_row']}: {exc}")
            continue

        try:
            human_priority = normalize_priority_label(row["priority_level"])
        except ValueError as exc:
            errors.append(f"Excel row {row['excel_row']}: {exc}")
            continue

        records.append(
            {
                "test_id": test_id,
                "title": source_row.get("title") or "",
                "description": source_description,
                "human_category_raw": row["human_label"],
                "human_category_normalized": human_category,
                "human_priority_raw": row["priority_level"],
                "human_priority_normalized": human_priority,
            }
        )

    if errors:
        preview = "\n".join(errors[:20])
        extra = "" if len(errors) <= 20 else f"\n... and {len(errors) - 20} more"
        raise ValueError(f"Owner-label validation failed:\n{preview}{extra}")

    if len(records) != 197:
        raise ValueError(f"Expected 197 owner-labeled records, found {len(records)}")

    return records


def build_export_records(records: list[dict]) -> list[ExportRecord]:
    export_records = []
    for index, row in enumerate(records, start=1):
        category_result = classify_issue_category(f"{row['title']} {row['description']}")
        priority_result = predict_priority(row["description"])
        if not priority_result.get("available") or not priority_result.get("priority"):
            raise ValueError(f"Priority model unavailable for test_id '{row['test_id']}'")

        system_category = normalize_category_label(category_result["category"])
        system_priority = normalize_priority_label(priority_result["priority"])
        confidence = float(priority_result["confidence"])
        export_records.append(
            ExportRecord(
                record_code=f"R{index:03d}",
                test_id=row["test_id"],
                sanitized_title=sanitize_text(row["title"]),
                sanitized_description=sanitize_text(row["description"]),
                human_category_raw=row["human_category_raw"],
                human_category_normalized=row["human_category_normalized"],
                system_category=system_category,
                category_correct=row["human_category_normalized"] == system_category,
                keyword_matched=bool(category_result.get("matched_keywords")),
                human_priority_raw=row["human_priority_raw"],
                human_priority_normalized=row["human_priority_normalized"],
                system_priority=system_priority,
                priority_confidence=confidence,
                priority_correct=row["human_priority_normalized"] == system_priority,
            )
        )
    return export_records


def _safe_div(numerator: int | float, denominator: int | float) -> float:
    return round(float(numerator) / float(denominator), 4) if denominator else 0.0


def calculate_multiclass_metrics(actual: list[str], predicted: list[str], label_order: list[str]) -> dict:
    confusion = {
        actual_label: {predicted_label: 0 for predicted_label in label_order}
        for actual_label in label_order
    }
    for actual_label, predicted_label in zip(actual, predicted):
        confusion[actual_label][predicted_label] += 1

    per_class = {}
    for label in label_order:
        tp = confusion[label][label]
        fp = sum(confusion[other][label] for other in label_order if other != label)
        fn = sum(confusion[label][other] for other in label_order if other != label)
        precision = _safe_div(tp, tp + fp)
        recall = _safe_div(tp, tp + fn)
        f1_score = round((2 * precision * recall) / (precision + recall), 4) if (precision + recall) else 0.0
        per_class[label] = {
            "precision": precision,
            "recall": recall,
            "f1_score": round(f1_score, 4),
            "support": sum(confusion[label].values()),
        }

    accuracy = _safe_div(sum(1 for a, p in zip(actual, predicted) if a == p), len(actual))
    macro_precision = round(mean(item["precision"] for item in per_class.values()), 4)
    macro_recall = round(mean(item["recall"] for item in per_class.values()), 4)
    macro_f1 = round(mean(item["f1_score"] for item in per_class.values()), 4)

    return {
        "total_records": len(actual),
        "correct_predictions": sum(1 for a, p in zip(actual, predicted) if a == p),
        "incorrect_predictions": sum(1 for a, p in zip(actual, predicted) if a != p),
        "accuracy": accuracy,
        "macro_precision": macro_precision,
        "macro_recall": macro_recall,
        "macro_f1": macro_f1,
        "per_class": per_class,
        "confusion_matrix": confusion,
    }


def compute_category_summary(export_records: list[ExportRecord]) -> dict:
    actual = [record.human_category_normalized for record in export_records]
    predicted = [record.system_category for record in export_records]
    metrics = calculate_multiclass_metrics(actual, predicted, CATEGORY_ORDER)
    keyword_match_count = sum(1 for record in export_records if record.keyword_matched)
    metrics["keyword_match_coverage"] = _safe_div(keyword_match_count, len(export_records))
    metrics["keyword_match_count"] = keyword_match_count
    metrics["unmatched_rate"] = _safe_div(len(export_records) - keyword_match_count, len(export_records))
    metrics["unmatched_count"] = len(export_records) - keyword_match_count
    return metrics


def compute_priority_summary(export_records: list[ExportRecord]) -> dict:
    actual = [record.human_priority_normalized for record in export_records]
    predicted = [record.system_priority for record in export_records]
    metrics = calculate_multiclass_metrics(actual, predicted, PRIORITY_ORDER)
    metrics["human_distribution"] = dict(Counter(actual))
    metrics["system_distribution"] = dict(Counter(predicted))
    metrics["average_confidence"] = round(mean(record.priority_confidence for record in export_records), 4)
    return metrics


def detect_identifier_leaks(export_records: Iterable[ExportRecord]) -> list[str]:
    leaks = []
    patterns = [
        ("email", EMAIL_RE),
        ("phone", PHONE_RE),
        ("unit", re.compile(r"\bUnit\s+[A-Z0-9-]+\b")),
        ("long_number", re.compile(r"(?<!\[Redacted )(?<!\d)\d{3,}(?!\d)")),
        ("payment", re.compile(r"(gcash|maya|account|payment|reference)", re.IGNORECASE)),
    ]
    for record in export_records:
        for label, pattern in patterns:
            for field_name, value in (
                ("sanitized_title", record.sanitized_title),
                ("sanitized_description", record.sanitized_description),
            ):
                if pattern.search(value):
                    leaks.append(f"{record.record_code} {field_name} contains potential {label} identifier")
    return leaks


def build_csv_rows(export_records: list[ExportRecord]) -> list[dict]:
    rows = []
    for record in export_records:
        rows.append(
            {
                "record_code": record.record_code,
                "sanitized_title": record.sanitized_title,
                "sanitized_description": record.sanitized_description,
                "human_category": title_case_category(record.human_category_normalized),
                "system_category": title_case_category(record.system_category),
                "category_correct": "Yes" if record.category_correct else "No",
                "keyword_matched": "Yes" if record.keyword_matched else "No",
                "human_priority": title_case_priority(record.human_priority_normalized),
                "system_priority": title_case_priority(record.system_priority),
                "priority_confidence": f"{record.priority_confidence * 100:.2f}%",
                "priority_correct": "Yes" if record.priority_correct else "No",
            }
        )
    return rows


def write_csv_export(path: Path, rows: list[dict]) -> None:
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=CSV_COLUMNS)
        writer.writeheader()
        writer.writerows(rows)


def _write_title(worksheet, title: str) -> None:
    worksheet["A1"] = title
    worksheet["A1"].font = Font(bold=True, size=14)


def _append_key_value_rows(worksheet, start_row: int, items: list[tuple[str, object]]) -> int:
    row_index = start_row
    for label, value in items:
        worksheet.cell(row=row_index, column=1, value=label)
        worksheet.cell(row=row_index, column=2, value=value)
        row_index += 1
    return row_index


def write_excel_export(
    path: Path,
    rows: list[dict],
    category_summary: dict,
    priority_summary: dict,
    validation_rows: list[tuple[str, object]],
) -> None:
    workbook = Workbook()

    details_sheet = workbook.active
    details_sheet.title = "Detailed Records"
    _write_title(details_sheet, "Maintenance Evaluation - Detailed Records")
    details_sheet.append([])
    details_sheet.append(CSV_COLUMNS)
    for row in rows:
        details_sheet.append([row[column] for column in CSV_COLUMNS])
    details_sheet.freeze_panes = "A4"

    category_sheet = workbook.create_sheet("Category Summary")
    _write_title(category_sheet, "Category Summary")
    row_index = _append_key_value_rows(
        category_sheet,
        3,
        [
            ("Total records", category_summary["total_records"]),
            ("Correct predictions", category_summary["correct_predictions"]),
            ("Incorrect predictions", category_summary["incorrect_predictions"]),
            ("Accuracy", category_summary["accuracy"]),
            ("Macro precision", category_summary["macro_precision"]),
            ("Macro recall", category_summary["macro_recall"]),
            ("Macro F1-score", category_summary["macro_f1"]),
            ("Keyword match coverage", category_summary["keyword_match_coverage"]),
            ("Unmatched rate", category_summary["unmatched_rate"]),
        ],
    )
    row_index += 1
    category_sheet.cell(row=row_index, column=1, value="Category")
    category_sheet.cell(row=row_index, column=2, value="Precision")
    category_sheet.cell(row=row_index, column=3, value="Recall")
    category_sheet.cell(row=row_index, column=4, value="F1-score")
    category_sheet.cell(row=row_index, column=5, value="Support")
    row_index += 1
    for label in CATEGORY_ORDER:
        metrics = category_summary["per_class"][label]
        category_sheet.cell(row=row_index, column=1, value=CATEGORY_DISPLAY[label])
        category_sheet.cell(row=row_index, column=2, value=metrics["precision"])
        category_sheet.cell(row=row_index, column=3, value=metrics["recall"])
        category_sheet.cell(row=row_index, column=4, value=metrics["f1_score"])
        category_sheet.cell(row=row_index, column=5, value=metrics["support"])
        row_index += 1

    category_confusion_sheet = workbook.create_sheet("Category Confusion Matrix")
    _write_title(category_confusion_sheet, "Category Confusion Matrix")
    category_confusion_sheet.append([])
    category_confusion_sheet.append(["Actual \\ Predicted"] + [CATEGORY_DISPLAY[label] for label in CATEGORY_ORDER])
    for actual_label in CATEGORY_ORDER:
        category_confusion_sheet.append(
            [CATEGORY_DISPLAY[actual_label]]
            + [category_summary["confusion_matrix"][actual_label][predicted_label] for predicted_label in CATEGORY_ORDER]
        )

    priority_sheet = workbook.create_sheet("Priority Summary")
    _write_title(priority_sheet, "Priority Summary")
    row_index = _append_key_value_rows(
        priority_sheet,
        3,
        [
            ("Total records", priority_summary["total_records"]),
            ("Correct predictions", priority_summary["correct_predictions"]),
            ("Incorrect predictions", priority_summary["incorrect_predictions"]),
            ("Accuracy", priority_summary["accuracy"]),
            ("Macro precision", priority_summary["macro_precision"]),
            ("Macro recall", priority_summary["macro_recall"]),
            ("Macro F1-score", priority_summary["macro_f1"]),
            ("Average confidence", priority_summary["average_confidence"]),
        ],
    )
    row_index += 1
    priority_sheet.cell(row=row_index, column=1, value="Priority")
    priority_sheet.cell(row=row_index, column=2, value="Precision")
    priority_sheet.cell(row=row_index, column=3, value="Recall")
    priority_sheet.cell(row=row_index, column=4, value="F1-score")
    priority_sheet.cell(row=row_index, column=5, value="Support")
    priority_sheet.cell(row=row_index, column=6, value="Human count")
    priority_sheet.cell(row=row_index, column=7, value="System count")
    row_index += 1
    for label in PRIORITY_ORDER:
        metrics = priority_summary["per_class"][label]
        priority_sheet.cell(row=row_index, column=1, value=PRIORITY_DISPLAY[label])
        priority_sheet.cell(row=row_index, column=2, value=metrics["precision"])
        priority_sheet.cell(row=row_index, column=3, value=metrics["recall"])
        priority_sheet.cell(row=row_index, column=4, value=metrics["f1_score"])
        priority_sheet.cell(row=row_index, column=5, value=metrics["support"])
        priority_sheet.cell(row=row_index, column=6, value=priority_summary["human_distribution"].get(label, 0))
        priority_sheet.cell(row=row_index, column=7, value=priority_summary["system_distribution"].get(label, 0))
        row_index += 1

    priority_confusion_sheet = workbook.create_sheet("Priority Confusion Matrix")
    _write_title(priority_confusion_sheet, "Priority Confusion Matrix")
    priority_confusion_sheet.append([])
    priority_confusion_sheet.append(["Actual \\ Predicted"] + [PRIORITY_DISPLAY[label] for label in PRIORITY_ORDER])
    for actual_label in PRIORITY_ORDER:
        priority_confusion_sheet.append(
            [PRIORITY_DISPLAY[actual_label]]
            + [priority_summary["confusion_matrix"][actual_label][predicted_label] for predicted_label in PRIORITY_ORDER]
        )

    validation_sheet = workbook.create_sheet("Data Validation")
    _write_title(validation_sheet, "Data Validation")
    validation_sheet.append([])
    for label, value in validation_rows:
        validation_sheet.append([label, value])

    for worksheet in workbook.worksheets:
        for column_cells in worksheet.columns:
            values = [len(str(cell.value)) for cell in column_cells if cell.value is not None]
            width = max(values, default=10) + 2
            worksheet.column_dimensions[column_cells[0].column_letter].width = min(width, 48)
        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def compare_expected_metrics(category_summary: dict, priority_summary: dict) -> list[str]:
    differences = []
    category_checks = {
        "total_records": category_summary["total_records"],
        "correct": category_summary["correct_predictions"],
        "incorrect": category_summary["incorrect_predictions"],
        "accuracy": category_summary["accuracy"],
        "macro_precision": category_summary["macro_precision"],
        "macro_recall": category_summary["macro_recall"],
        "macro_f1": category_summary["macro_f1"],
        "keyword_match_coverage": category_summary["keyword_match_coverage"],
        "unmatched_rate": category_summary["unmatched_rate"],
    }
    for key, expected in EXPECTED_CATEGORY_METRICS.items():
        actual = category_checks[key]
        if actual != expected:
            differences.append(f"Category metric mismatch for {key}: expected {expected}, actual {actual}")

    priority_checks = {
        "total_records": priority_summary["total_records"],
        "accuracy": priority_summary["accuracy"],
        "macro_f1": priority_summary["macro_f1"],
    }
    for key, expected in EXPECTED_PRIORITY_METRICS.items():
        if key == "human_distribution":
            actual_distribution = priority_summary["human_distribution"]
            if actual_distribution != expected:
                differences.append(
                    f"Priority human distribution mismatch: expected {expected}, actual {actual_distribution}"
                )
            continue
        actual = priority_checks[key]
        if actual != expected:
            differences.append(f"Priority metric mismatch for {key}: expected {expected}, actual {actual}")
    return differences


def build_summary_text(
    workbook_path: Path,
    source_csv_path: Path,
    csv_path: Path,
    xlsx_path: Path,
    category_summary: dict,
    priority_summary: dict,
    differences: list[str],
) -> str:
    lines = [
        "RealEstate360+ Maintenance Evaluation Export",
        "",
        f"Human-label source workbook: {workbook_path}",
        f"Maintenance source dataset: {source_csv_path}",
        "MaintenanceRequest model: maintenance/models.py -> MaintenanceRequest",
        "Category function: accounts/ml/maintenance_nlp.py -> classify_issue_category(text)",
        "Priority function: accounts/ml/maintenance_nlp.py -> predict_priority(text)",
        "Category preprocessing: _clean_text(text), then keyword matching on classify_issue_category(f\"{title} {description}\")",
        "Priority preprocessing: _clean_text(text), then model.predict_proba(cleaned_description)",
        "Priority confidence source: top predicted probability returned by predict_priority(...), stored as a 0.0-1.0 float and exported as a percentage with two decimals",
        "",
        f"Generated CSV: {csv_path}",
        f"Generated Excel workbook: {xlsx_path}",
        "",
        "Category summary:",
        json.dumps(category_summary, indent=2),
        "",
        "Priority summary:",
        json.dumps(priority_summary, indent=2),
        "",
    ]
    if differences:
        lines.extend(["Discrepancies against expected values:"] + differences)
    else:
        lines.append("All calculated metrics matched the expected values exactly.")
    return "\n".join(lines) + "\n"


def export_maintenance_evaluation(
    workbook_path: Path = DEFAULT_OWNER_WORKBOOK,
    source_csv_path: Path = DEFAULT_SOURCE_DATASET,
    reports_dir: Path = DEFAULT_REPORTS_DIR,
) -> dict:
    owner_records = load_owner_label_records(workbook_path=workbook_path, source_csv_path=source_csv_path)
    export_records = build_export_records(owner_records)

    identifier_leaks = detect_identifier_leaks(export_records)
    if identifier_leaks:
        preview = "\n".join(identifier_leaks[:20])
        extra = "" if len(identifier_leaks) <= 20 else f"\n... and {len(identifier_leaks) - 20} more"
        raise ValueError(f"Sanitized export still contains potential identifiers:\n{preview}{extra}")

    csv_rows = build_csv_rows(export_records)
    category_summary = compute_category_summary(export_records)
    priority_summary = compute_priority_summary(export_records)
    metric_differences = compare_expected_metrics(category_summary, priority_summary)

    reports_dir.mkdir(parents=True, exist_ok=True)
    csv_path = reports_dir / "maintenance_evaluation_197.csv"
    xlsx_path = reports_dir / "maintenance_evaluation_197.xlsx"
    summary_path = reports_dir / "maintenance_evaluation_summary.txt"

    validation_rows = [
        ("Expected record count", 197),
        ("Actual record count", len(export_records)),
        ("Unique test IDs", len({record.test_id for record in export_records})),
        ("CSV columns", ", ".join(CSV_COLUMNS)),
        ("Workbook source", str(workbook_path)),
        ("Maintenance source dataset", str(source_csv_path)),
        ("Category function", "accounts.ml.maintenance_nlp.classify_issue_category"),
        ("Priority function", "accounts.ml.maintenance_nlp.predict_priority"),
        ("Category input flow", 'classify_issue_category(f"{title} {description}")'),
        ("Priority input flow", "predict_priority(description)"),
        ("Priority confidence format", "top predicted probability formatted as XX.XX%"),
        ("Potential identifier leaks", "None"),
        ("Metric discrepancies", "None" if not metric_differences else "See summary file"),
    ]

    write_csv_export(csv_path, csv_rows)
    write_excel_export(xlsx_path, csv_rows, category_summary, priority_summary, validation_rows)
    summary_text = build_summary_text(
        workbook_path=workbook_path,
        source_csv_path=source_csv_path,
        csv_path=csv_path,
        xlsx_path=xlsx_path,
        category_summary=category_summary,
        priority_summary=priority_summary,
        differences=metric_differences,
    )
    summary_path.write_text(summary_text, encoding="utf-8")

    return {
        "workbook_path": str(workbook_path),
        "source_csv_path": str(source_csv_path),
        "csv_path": str(csv_path),
        "xlsx_path": str(xlsx_path),
        "summary_path": str(summary_path),
        "csv_rows": csv_rows,
        "export_records": export_records,
        "category_summary": category_summary,
        "priority_summary": priority_summary,
        "metric_differences": metric_differences,
    }
